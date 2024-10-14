#!D:/Code/python
# -*- coding: utf-8 -*-
# @Time : 2021/7/28 16:37
# @Author : chaunwen.peng
# @File : main.py
# @Software: PyCharm
import time

from .PTZControl import PTZControl
from .TrackControl import TrackControl
from .AK.akAction import AkAction
from .AKMeasureTool import getValue

from openpyxl import load_workbook


def op_toExcel(data, file_name, count=0):  # openpyxl库储存数据到excel
    wb = load_workbook(file_name)
    ws = wb['all_test']  # 创建子表
    for i in range(len(data)):
        ws.cell(row=16 + count, column=1 + i, value=data[i])
    wb.save(file_name)


def move(file_name):
    dis_lis = [1060, 1590, 2120, 2660, 3170]
    angle_dic = {
        "2120": [325, 335, 345, 0, 15, 25, 35],
        "2660": [330, 340, 350, 0, 10, 20, 30]
    }
    # for dis in dis_lis:
    #     time.sleep(5)
    #     ret = track_con.move_to(3500, dis, 180)
    angle_lis = angle_dic["2120"] if 1060 >= 2120 else angle_dic["2660"]
    for count, angle in enumerate([10, 20, 30]):
        # for count, angle in enumerate(angle_lis):
        ptz_con.turn_to_angle(angle=angle)
        AkAction().doAk()
        time.sleep(3)
        # ret = getValue()
        # ret = [i for i in ret]
        # ret.pop(5)
        # retry = 0
        # while len(set(ret)) == 1 and retry < 5:
        #     retry += 1
        #     ret = getValue()
        #     ret = [i for i in ret]
        #     ret.pop(5)
        # ret.insert(0, angle)
        # print(ret)
        # op_toExcel(ret, file_name, count=count)
    ptz_con.turn_to_angle(axis="H", angle=10)
    AkAction().doAk()
    time.sleep(3)
    # ret = getValue()
    # ret = [i for i in ret]
    # ret.pop(5)
    # retry = 0
    # while len(set(ret)) == 1 and retry < 5:
    #     retry += 1
    #     ret = getValue()
    #     ret = [i for i in ret]
    #     ret.pop(5)
    # ret.insert(0, 10)
    # print(ret)
    # op_toExcel(ret, file_name, count=7)


def stress_test_track(count=0):
    while count < 10000:
        track_con.move_to(0, 2160)
        time.sleep(2)
        ret = getValue()
        ret = [i for i in ret]
        ret.pop(5)
        print(ret)
        op_toExcel(ret, file_name, count=count)

        track_con.move_to(500, 2160)
        time.sleep(2)
        track_con.move_to(500, 3160)
        time.sleep(2)
        track_con.move_to(0, 3160)
        count += 1


def stress_test_angle(count=0):
    while count < 10000:
        print("------------------------第%d次测试-------------------------" % (count + 1))
        ptz_con.reset()
        time.sleep(2)
        ret = getValue()
        ret = [i for i in ret]
        ret.pop(5)
        # print(ret)
        op_toExcel(ret, file_name, count=count)

        ptz_con.turn_to_angle(angle=25)

        time.sleep(2)
        ptz_con.reset()

        ret = getValue()
        ret = [i for i in ret]
        ret.pop(5)
        print(ret)
        op_toExcel(ret, file_name, count=count)

        ptz_con.turn_to_angle(angle=335)

        count += 1


def stress_test_angle2(count):
    ptz_con.reset()
    time.sleep(2)
    flag = True if abs(track_con.get_cur()[-1] - 2789) <= 3 and not track_con.get_cur()[0] else False
    if flag:
        AkAction().doAk3()
        ret = getValue()
        ret = [i for i in ret]
        ret.pop(5)
        # print(ret)
        op_toExcel(ret, file_name, count=count)

    ptz_con.turn_to_angle(angle=25)

    time.sleep(5)
    ptz_con.reset()

    if flag:
        AkAction().doAk3()
        ret = getValue()
        ret = [i for i in ret]
        ret.pop(5)
        # print(ret)
        op_toExcel(ret, file_name, count=count + 1)

    ptz_con.turn_to_angle(angle=335)
    ptz_con.reset()
    time.sleep(5)
    if flag:
        AkAction().doAk3()
        ret = getValue()
        ret = [i for i in ret]
        ret.pop(5)
        # print(ret)
        op_toExcel(ret, file_name, count=count + 2)


def track_angle(count=0):
    while count < 10000:
        print("------------------------第%d次测试-------------------------" % (count + 1))
        print("开始移动到（0， 2660）")
        track_con.move_to(0, 2660)
        time.sleep(2)
        stress_test_angle2(count)

        print("开始移动到（1000， 2660）")
        track_con.move_to(1000, 2660)
        time.sleep(2)
        # stress_test_angle2(count)

        print("开始移动到（1000， 3660）")
        track_con.move_to(1000, 3660)
        time.sleep(2)
        # stress_test_angle2(count)

        print("开始移动到（0， 3660）")
        time.sleep(2)
        track_con.move_to(0, 3660)
        time.sleep(2)
        # stress_test_angle2(count)
        count += 3


def test_new_angle():
    angle_dic = {
        "2120": [325, 335, 345, 0, 15, 25, 35],
        "2660": [330, 340, 350, 0, 10, 20, 30]
    }
    for count, angle in enumerate(angle_dic):
        ptz_con.turn_to_angle(angle=angle)
        time.sleep(3)
        ret = getValue()
        ret = [i for i in ret]
        ret.pop(5)
        retry = 0
        while len(set(ret)) == 1 and retry < 5:
            retry += 1
            ret = getValue()
            ret = [i for i in ret]
            ret.pop(5)
        ret.insert(0, angle)
        print(ret)
        op_toExcel(ret, file_name, count=count)


if __name__ == "__main__":
    ptz_con = PTZControl()
    track_con = TrackControl()
    # print(track_con.move_to(3500, 2660, 180))
    # ptz_con.turn_to_angle(angle=5)
    # AkAction().doAk()
    # time.sleep(3)
    # ret = getValue()
    # print("ret: %s" % ret)
    # track_con.reset()
    # ptz_con.reset()

    # track_con.move_to(3500, 2160, 180)
    file_name = 'AK_test.xlsx'
    # move(file_name)
    # data = [2140.0, 1731.0, 1333.0, 1279.0, 2349.0, 80.45841979980469, 97.95682525634766, 100.22189331054688,
    #         81.36286163330078, 10.221893310546875, -1.0]

    # op_toExcel(data, file_name)
    # stress_test_track()
    # stress_test_angle(count=190)
    track_angle()
